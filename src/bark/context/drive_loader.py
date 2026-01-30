"""Google Drive loader for fetching content."""

import io
import logging
import os
import hashlib
from dataclasses import dataclass, field
from typing import Any

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from google.oauth2 import service_account
import openpyxl
import pytesseract
from pdf2image import convert_from_bytes

logger = logging.getLogger(__name__)

# If modifying these scopes, delete the file token.json.
SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]


@dataclass
class DriveChunk:
    """A chunk of Drive content."""

    id: str
    content: str
    metadata: dict[str, Any] = field(default_factory=dict)


import json

# ... imports ...


class DriveLoader:
    """Loads and parses content from Google Drive."""

    def __init__(
        self,
        credentials_file: str | None = "credentials.json",
        credentials_json: str | None = None,
        token_json: str | None = None,
        folder_id: str | None = None,
        exclude_folder_ids: list[str] | None = None,
        chunk_size: int = 500,
        token_file: str = "token.json",
    ) -> None:
        """Initialize the Drive loader.

        Args:
            credentials_file: Path to credentials.json
            credentials_json: Content of credentials.json (optional override)
            token_json: Content of token.json (optional override)
            folder_id: Optional folder ID to scope search
            exclude_folder_ids: Optional list of folder IDs to exclude from indexing
            chunk_size: Target size for content chunks (in words)
            token_file: Path to store user access tokens
        """
        self.credentials_file = credentials_file
        self.credentials_json = credentials_json
        self.token_json = token_json
        self.folder_id = folder_id
        self.exclude_folder_ids = set(exclude_folder_ids or [])
        self.chunk_size = chunk_size
        self.token_file = token_file
        self._service: Any | None = None

    def _get_service(self) -> Any:
        """Get or create the Drive service."""
        if self._service:
            return self._service

        creds = None
        
        # 1. Try Service Account from JSON string
        if self.credentials_json:
            try:
                info = json.loads(self.credentials_json)
                creds = service_account.Credentials.from_service_account_info(
                    info, scopes=SCOPES
                )
                logger.info("Loaded Service Account credentials from JSON string")
            except Exception as e:
                logger.warning(f"Failed to load Service Account from JSON string: {e}")

        # 2. Try Service Account from File
        if not creds and self.credentials_file and os.path.exists(self.credentials_file):
            try:
                # Try loading as service account first
                creds = service_account.Credentials.from_service_account_file(
                    self.credentials_file, scopes=SCOPES
                )
                logger.info("Loaded Service Account credentials from file")
            except ValueError:
                # Fallback to user credentials flow
                pass

        if not creds:
             # 3. Try User Token from JSON string
            if self.token_json:
                try:
                    info = json.loads(self.token_json)
                    creds = Credentials.from_authorized_user_info(info, SCOPES)
                    logger.info("Loaded User credentials from JSON string")
                except Exception as e:
                    logger.warning(f"Failed to load User credentials from JSON string: {e}")

            # 4. Try User Token from File
            if (not creds or not creds.valid) and os.path.exists(self.token_file):
                creds = Credentials.from_authorized_user_file(self.token_file, SCOPES)
            
            # If there are no (valid) credentials available, let the user log in.
            if not creds or not creds.valid:
                if creds and creds.expired and creds.refresh_token:
                    creds.refresh(Request())
                else:
                    if not self.credentials_file or not os.path.exists(self.credentials_file):
                         raise FileNotFoundError(f"Credentials file not found: {self.credentials_file}")
                    
                    flow = InstalledAppFlow.from_client_secrets_file(
                        self.credentials_file, SCOPES
                    )
                    creds = flow.run_local_server(port=0)
                
                # Save the credentials for the next run (only if file path is set)
                if self.token_file:
                    with open(self.token_file, "w") as token:
                        token.write(creds.to_json())
                logger.info("Loaded User credentials")

        self._service = build("drive", "v3", credentials=creds)
        return self._service

    def load(self, file_ids: list[str] | None = None) -> list[DriveChunk]:
        """Fetch and parse Drive files.
        
        Args:
            file_ids: Optional list of file IDs to fetch. If None, fetches all files in folder.
            
        Returns:
            List of Drive chunks
        """
        chunks: list[DriveChunk] = []
        service = self._get_service()

        try:
            files_to_process = []
            if file_ids:
                for file_id in file_ids:
                    try:
                        file = service.files().get(
                            fileId=file_id, 
                            fields="id, name, mimeType, modifiedTime, webViewLink",
                            supportsAllDrives=True,
                        ).execute()
                        files_to_process.append(file)
                    except Exception as e:
                        logger.warning(f"Failed to retrieve file {file_id}: {e}")
            else:
                files_to_process = self._list_files(service)

            logger.warning(f"Processing {len(files_to_process)} Drive files")

            for i, file in enumerate(files_to_process, 1):
                if i % 5 == 0:
                    logger.warning(f"Processing file {i}/{len(files_to_process)}: {file.get('name', 'Untitled')}")
                
                file_chunks = self._process_file(service, file)
                chunks.extend(file_chunks)

            return chunks

        except Exception as e:
            logger.error(f"Failed to load Drive content: {e}")
            raise

    def fetch_file_metadata(self) -> dict[str, str]:
        """Fetch metadata for all accessible files.

        Returns:
            Dictionary mapping file ID to modifiedTime
        """
        service = self._get_service()
        metadata = {}

        try:
            files = self._list_files(service)
            for file in files:
                metadata[file["id"]] = file.get("modifiedTime", "")
        except Exception as e:
            logger.error(f"Failed to fetch Drive metadata: {e}")
            raise

        return metadata

    def search(self, query: str, max_results: int = 5) -> list[dict]:
        """Search Google Drive files using the native fullText search API.

        Args:
            query: Search query to match against file content and names
            max_results: Maximum number of results to return

        Returns:
            List of search results with file info and content preview
        """
        results = []
        service = self._get_service()
        
        try:
            # Build query for fullText and name search
            # Escape quotes in query
            escaped_query = query.replace("'", "\\'")
            
            # Search for files containing the query text or matching name
            # Ignore self.folder_id for live search to be more useful
            search_query = f"(fullText contains '{escaped_query}' or name contains '{escaped_query}') and trashed = false"

            # Add mime type filter for supported types
            mime_types = [
                "application/vnd.google-apps.document",
                "application/vnd.google-apps.spreadsheet",
                "application/vnd.google-apps.presentation",
                "text/plain",
                "text/markdown",
                "application/pdf",
                "application/vnd.google-apps.folder",
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "application/vnd.openxmlformats-officedocument.presentationml.presentation",
            ]
            mime_query = " or ".join(f"mimeType = '{mt}'" for mt in mime_types)
            search_query += f" and ({mime_query})"
            
            logger.info(f"Drive search query: {search_query}")
            
            # Use corpora='allDrives' to search everything accessible, including shared drives
            response = service.files().list(
                q=search_query,
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
                corpora="allDrives",
                fields="files(id, name, mimeType, modifiedTime, webViewLink)",
                pageSize=max_results * 2,
            ).execute()
            
            files = response.get("files", [])
            logger.info(f"Drive search found {len(files)} matches")
            
            for file in files:
                file_id = file["id"]
                name = file.get("name", "Untitled")
                mime_type = file.get("mimeType", "")
                
                # Skip files with "resume" in the title
                if "resume" in name.lower():
                    continue
                
                # Fetch content preview
                try:
                    content = ""
                    if mime_type == "application/vnd.google-apps.folder":
                        content = "(This is a folder. You can list its contents using other tools if available, or search for files inside it.)"
                    elif mime_type == "application/vnd.google-apps.document":
                        content = self._export_gdoc(service, file_id)
                    elif mime_type == "application/vnd.google-apps.spreadsheet":
                        content = self._export_gsheet(service, file_id)
                    elif mime_type == "application/vnd.google-apps.presentation":
                        content = self._export_gslide(service, file_id)
                    elif mime_type in ["text/plain", "text/markdown"]:
                        content = self._download_file(service, file_id)
                    elif mime_type == "application/pdf":
                        # For preview in search, we might want to skip heavy OCR or just get metadata
                        content = "(PDF file - use indexing/memory tools for full content extract)"
                    
                    # Truncate content for preview
                    words = content.split()
                    preview = " ".join(words[:200]) + ("..." if len(words) > 200 else "")
                    
                    results.append({
                        "title": f"[FOLDER] {name}" if mime_type == "application/vnd.google-apps.folder" else name,
                        "url": file.get("webViewLink", ""),
                        "file_id": file_id,
                        "content": preview,
                        "mime_type": mime_type,
                        "last_edited": file.get("modifiedTime", ""),
                    })
                except Exception as e:
                    logger.warning(f"Could not fetch content for file {name}: {e}")
                    results.append({
                        "title": name,
                        "url": file.get("webViewLink", ""),
                        "file_id": file_id,
                        "content": "(Content unavailable)",
                        "mime_type": mime_type,
                        "last_edited": file.get("modifiedTime", ""),
                    })
                    
                if len(results) >= max_results:
                    break
                    
        except Exception as e:
            logger.error(f"Drive search failed: {e}")
            raise
            
        return results

    def _list_files(self, service: Any) -> list[dict[str, Any]]:
        """List files in the configured scope (recursively if folder_id is set)."""
        all_files = []
        
        # If no folder_id, search everything (flat search is effectively recursive for the user's view)
        if not self.folder_id:
            return self._perform_search(service, parent_id=None)

        # If folder_id is set, crawl recursively
        logger.warning(f"Starting recursive crawl of folder: {self.folder_id}")
        folders_to_process = [self.folder_id]
        processed_folders = set()

        while folders_to_process:
            current_folder_id = folders_to_process.pop(0)
            
            if current_folder_id in processed_folders:
                continue
            
            # Skip excluded folders (check here too in case it's in the initial queue)
            if current_folder_id in self.exclude_folder_ids:
                logger.warning(f"Skipping excluded folder: {current_folder_id}")
                continue
                
            processed_folders.add(current_folder_id)
            
            if len(processed_folders) % 5 == 0:
                 logger.warning(f"Crawled {len(processed_folders)} folders...")

            # Fetch items in this folder (both content and subfolders)
            items = self._perform_search(service, parent_id=current_folder_id, include_folders=True)
            
            for item in items:
                if item.get("mimeType") == "application/vnd.google-apps.folder":
                    # Skip excluded folders
                    if item["id"] not in self.exclude_folder_ids:
                        folders_to_process.append(item["id"])
                    else:
                        logger.warning(f"Skipping excluded folder: {item.get('name', item['id'])}")
                else:
                    all_files.append(item)
                    
        return all_files

    def _perform_search(self, service: Any, parent_id: str | None, include_folders: bool = False) -> list[dict[str, Any]]:
        """Helper to safely search a specific scope."""
        files = []
        page_token = None
        
        query_parts = ["trashed = false"]
        
        if parent_id:
            query_parts.append(f"'{parent_id}' in parents")
        
        mime_types = [
            "application/vnd.google-apps.document",
            "application/vnd.google-apps.spreadsheet",
            "application/vnd.google-apps.presentation",
            "text/plain",
            "text/markdown",
            "application/pdf",
        ]
        if include_folders:
            mime_types.append("application/vnd.google-apps.folder")
            
        mime_query = " or ".join(f"mimeType = '{mt}'" for mt in mime_types)
        query_parts.append(f"({mime_query})")
        
        query = " and ".join(query_parts)

        while True:
            try:
                response = service.files().list(
                    q=query,
                    spaces="drive",
                    fields="nextPageToken, files(id, name, mimeType, modifiedTime, webViewLink)",
                    pageToken=page_token,
                    supportsAllDrives=True,
                    includeItemsFromAllDrives=True,
                ).execute()
                
                files.extend(response.get("files", []))
                page_token = response.get("nextPageToken", None)
                if page_token is None:
                    break
            except Exception as e:
                logger.warning(f"Error searching in {parent_id}: {e}")
                break
                
        return files

    def _process_file(self, service: Any, file: dict[str, Any]) -> list[DriveChunk]:
        """Process a single file into chunks."""
        chunks = []
        file_id = file["id"]
        name = file.get("name", "Untitled")
        mime_type = file.get("mimeType", "")
        
        # Skip files with "resume" in the title
        if "resume" in name.lower():
            logger.warning(f"Skipping file with 'resume' in title: {name}")
            return []
        
        content = ""
        
        try:
            if mime_type == "application/vnd.google-apps.document":
                # Export Google Doc as text
                content = self._export_gdoc(service, file_id)
            elif mime_type == "application/vnd.google-apps.spreadsheet":
                # Export Google Sheet as CSV
                content = self._export_gsheet(service, file_id)
            elif mime_type == "application/vnd.google-apps.presentation":
                # Export Google Slides as text
                content = self._export_gslide(service, file_id)
            elif mime_type in ["text/plain", "text/markdown"]:
                # Download text file
                content = self._download_file(service, file_id)
            elif mime_type == "application/pdf":
                # Extract text from PDF via OCR
                content = self._extract_pdf_text(service, file_id)
            
            if not content.strip():
                return []

            text_chunks = self._split_into_chunks(content)
            
            for i, chunk_text in enumerate(text_chunks):
                chunk_id = self._generate_chunk_id(file_id, i)
                chunks.append(
                    DriveChunk(
                        id=chunk_id,
                        content=chunk_text,
                        metadata={
                            "page": name,
                            "source": f"drive/{file_id}",
                            "url": file.get("webViewLink", ""),
                            "source_type": "drive",
                            "last_edited_time": file.get("modifiedTime", ""),
                            "mime_type": mime_type,
                        },
                    )
                )

        except Exception as e:
            logger.warning(f"Failed to process file {name} ({file_id}): {e}")
            
        return chunks

    def _export_gdoc(self, service: Any, file_id: str) -> str:
        """Export a Google Doc to plain text."""
        request = service.files().export_media(
            fileId=file_id, mimeType="text/plain"
        )
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            
        return fh.getvalue().decode("utf-8")

    def _export_gsheet(self, service: Any, file_id: str) -> str:
        """Export a Google Sheet to text (parsing all sheets) in LLM-readable markdown format."""
        # Export as XLSX
        request = service.files().export_media(
            fileId=file_id, 
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
        
        # Parse with openpyxl
        fh.seek(0)
        wb = openpyxl.load_workbook(fh, read_only=True, data_only=True)
        
        output = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            output.append(f"## Sheet: {sheet_name}\n")
            
            # Collect all rows first to determine column count and detect header
            rows = []
            for row in ws.iter_rows(values_only=True):
                # Convert cells, replacing None with empty string
                row_text = [str(cell) if cell is not None else "" for cell in row]
                # Skip completely empty rows
                if any(cell.strip() for cell in row_text):
                    rows.append(row_text)
            
            if not rows:
                output.append("(Empty sheet)\n")
                continue
            
            # Determine max column count
            max_cols = max(len(row) for row in rows)
            
            # Pad rows to have consistent column count
            for i, row in enumerate(rows):
                while len(row) < max_cols:
                    rows[i].append("")
            
            # First row is treated as header
            header = rows[0]
            output.append("| " + " | ".join(header) + " |")
            output.append("| " + " | ".join(["---"] * max_cols) + " |")
            
            # Data rows with row numbers for reference
            for row_num, row in enumerate(rows[1:], start=2):
                output.append("| " + " | ".join(row) + " |")
            
            output.append("\n")
            
        return "\n".join(output)

    def _export_gslide(self, service: Any, file_id: str) -> str:
        """Export a Google Slide to plain text."""
        request = service.files().export_media(
            fileId=file_id, mimeType="text/plain"
        )
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            
        return fh.getvalue().decode("utf-8")

    def _download_file(self, service: Any, file_id: str) -> str:
        """Download a binary file (text)."""
        request = service.files().get_media(fileId=file_id)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
            
        return fh.getvalue().decode("utf-8")

    def _extract_pdf_text(self, service: Any, file_id: str) -> str:
        """Extract text from a PDF using Tesseract OCR."""
        # Download the PDF bytes
        request = service.files().get_media(fileId=file_id)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
        
        pdf_bytes = fh.getvalue()
        
        # Convert PDF pages to images
        images = convert_from_bytes(pdf_bytes)
        
        # OCR each page
        text_parts = []
        for i, image in enumerate(images):
            page_text = pytesseract.image_to_string(image)
            if page_text.strip():
                text_parts.append(f"--- Page {i + 1} ---\n{page_text}")
        
        return "\n\n".join(text_parts)

    def _split_into_chunks(self, text: str) -> list[str]:
        """Split text into chunks of approximately chunk_size words."""
        words = text.split()

        if len(words) <= self.chunk_size:
            return [text]

        chunks: list[str] = []
        current_chunk: list[str] = []

        for word in words:
            current_chunk.append(word)
            if len(current_chunk) >= self.chunk_size:
                chunks.append(" ".join(current_chunk))
                current_chunk = []

        if current_chunk:
            chunks.append(" ".join(current_chunk))

        return chunks

    def _generate_chunk_id(self, file_id: str, index: int) -> str:
        """Generate a unique ID for a chunk."""
        raw_id = f"drive:{file_id}:{index}"
        return hashlib.md5(raw_id.encode()).hexdigest()[:16]
